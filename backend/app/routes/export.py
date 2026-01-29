from flask import Blueprint, send_file, jsonify, request
from flask_jwt_extended import jwt_required
from datetime import datetime
from app.services.export_service import RoadmapExporter, PRDExporter
from app.services import CategoryService
from app.models import Category, Feature
from app import db
from openpyxl import load_workbook
from io import BytesIO

export_bp = Blueprint("export", __name__)


def parse_release_date(date_str):
    """Convert 'January 2024' format back to '2024-01' format"""
    if not date_str:
        return None
    try:
        # Try parsing as "Month Year" format
        date_obj = datetime.strptime(date_str, "%B %Y")
        return date_obj.strftime("%Y-%m")
    except (ValueError, TypeError):
        # Already in correct format or invalid
        return date_str


@export_bp.route("/export/roadmap/pptx", methods=["POST"])
@jwt_required()
def export_roadmap_pptx():
    """Export roadmap to PowerPoint presentation"""
    try:
        # Fetch all categories with features
        categories = Category.query.all()

        # Check if any features have release dates
        has_dates = any(
            feature.release_date for cat in categories for feature in cat.features
        )

        if not has_dates:
            return (
                jsonify(
                    {
                        "error": "No features with release dates found. Cannot generate roadmap."
                    }
                ),
                400,
            )

        # Generate PowerPoint
        exporter = RoadmapExporter(categories)
        pptx_buffer = exporter.generate_pptx()

        # Generate filename with current date
        filename = f'roadmap_{datetime.now().strftime("%Y-%m-%d")}.pptx'

        # Send file
        return send_file(
            pptx_buffer,
            mimetype="application/vnd.openxmlformats-officedocument.presentationml.presentation",
            as_attachment=True,
            download_name=filename,
        )

    except ValueError as e:
        # Handle specific errors from the exporter
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        # Handle unexpected errors
        return jsonify({"error": f"Failed to generate PowerPoint: {str(e)}"}), 500


@export_bp.route("/export/prd/excel", methods=["POST"])
@jwt_required()
def export_prd_excel():
    """Export all categories and features to Excel"""
    try:
        # Fetch all categories with features
        categories = Category.query.all()

        if not categories:
            return jsonify({"error": "No categories found. Cannot generate PRD."}), 400

        # Generate Excel
        exporter = PRDExporter(categories)
        excel_buffer = exporter.generate_excel()

        # Generate filename with current date
        filename = f'prd_{datetime.now().strftime("%Y-%m-%d")}.xlsx'

        # Send file
        return send_file(
            excel_buffer,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )

    except ValueError as e:
        # Handle specific errors from the exporter
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        # Handle unexpected errors
        return jsonify({"error": f"Failed to generate Excel: {str(e)}"}), 500


@export_bp.route("/export/prd/word", methods=["POST"])
@jwt_required()
def export_prd_word():
    """Export all categories and features to Word document"""
    try:
        # Fetch all categories with features
        categories = Category.query.all()

        if not categories:
            return jsonify({"error": "No categories found. Cannot generate PRD."}), 400

        # Generate Word document
        exporter = PRDExporter(categories)
        word_buffer = exporter.generate_word()

        # Generate filename with current date
        filename = f'prd_{datetime.now().strftime("%Y-%m-%d")}.docx'

        # Send file
        return send_file(
            word_buffer,
            mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            as_attachment=True,
            download_name=filename,
        )

    except ValueError as e:
        # Handle specific errors from the exporter
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        # Handle unexpected errors
        return jsonify({"error": f"Failed to generate Word document: {str(e)}"}), 500


@export_bp.route("/import/prd/excel", methods=["POST"])
@jwt_required()
def import_prd_excel():
    """Import categories and features from Excel file"""
    try:
        # Check if file is present
        if "file" not in request.files:
            return jsonify({"error": "No file provided"}), 400

        file = request.files["file"]

        if file.filename == "":
            return jsonify({"error": "No file selected"}), 400

        if not file.filename.endswith(".xlsx"):
            return jsonify({"error": "File must be an Excel file (.xlsx)"}), 400

        # Read the Excel file
        wb = load_workbook(filename=BytesIO(file.read()))

        # Skip the summary sheet, process category sheets
        imported_categories = 0
        imported_features = 0

        for sheet_name in wb.sheetnames:
            if sheet_name in ["Summary", "PRD Summary"]:
                continue

            sheet = wb[sheet_name]

            # Get category description from the sheet (should be in row 1)
            category_description = sheet.cell(row=1, column=1).value or ""
            if category_description.startswith("Category: "):
                category_description = category_description.replace("Category: ", "", 1)

            # Create or update category
            # Note: Import assumes default project (id=1) for backward compatibility
            category = Category.query.filter_by(name=sheet_name).first()
            if not category:
                category = CategoryService.create_category(
                    name=sheet_name, project_id=1, description=category_description
                )
                imported_categories += 1
            else:
                category.description = category_description
                db.session.add(category)

            # Read features from row 5 onwards (row 4 is header, row 5 is first data row)
            row_num = 5
            while True:
                feature_id = sheet.cell(row=row_num, column=1).value

                # Stop if we hit an empty row
                if not feature_id:
                    break

                title = sheet.cell(row=row_num, column=2).value or ""
                priority = sheet.cell(row=row_num, column=3).value or "Medium"
                description = sheet.cell(row=row_num, column=4).value or ""
                kpi = sheet.cell(row=row_num, column=5).value or ""
                customer_name = sheet.cell(row=row_num, column=6).value or ""
                engineering_comment = sheet.cell(row=row_num, column=7).value or ""
                engineering_signoff_str = (
                    sheet.cell(row=row_num, column=8).value or "No"
                )
                engineering_signoff = engineering_signoff_str == "Yes"
                engineering_complexity = sheet.cell(row=row_num, column=9).value or "M"
                release_date_val = sheet.cell(row=row_num, column=10).value

                # Parse release date
                release_date = None
                if release_date_val:
                    if isinstance(release_date_val, str):
                        release_date = parse_release_date(release_date_val)
                    else:
                        # If it's a datetime object, format it
                        try:
                            release_date = release_date_val.strftime("%Y-%m")
                        except (AttributeError, TypeError):
                            release_date = None

                # Check if feature already exists
                feature = Feature.query.filter_by(id=feature_id).first()
                if not feature:
                    # Create new feature
                    feature = Feature(
                        id=feature_id,
                        title=title,
                        priority=priority,
                        description=description,
                        kpi=kpi,
                        customer_name=customer_name,
                        engineering_comment=engineering_comment,
                        engineering_signoff=engineering_signoff,
                        engineering_complexity=engineering_complexity,
                        release_date=release_date,
                        category_id=category.id,
                    )
                    db.session.add(feature)
                    imported_features += 1
                else:
                    # Update existing feature
                    feature.title = title
                    feature.priority = priority
                    feature.description = description
                    feature.kpi = kpi
                    feature.customer_name = customer_name
                    feature.engineering_comment = engineering_comment
                    feature.engineering_signoff = engineering_signoff
                    feature.engineering_complexity = engineering_complexity
                    feature.release_date = release_date
                    feature.category_id = category.id

                row_num += 1

        # Commit all changes
        db.session.commit()

        return (
            jsonify(
                {
                    "message": "Import successful",
                    "categories_imported": imported_categories,
                    "features_imported": imported_features,
                }
            ),
            200,
        )

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Failed to import Excel file: {str(e)}"}), 500
