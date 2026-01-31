import pytest
from io import BytesIO
from openpyxl import load_workbook
from pptx import Presentation
from docx import Document
from app.models import Category, Feature


@pytest.mark.functional
class TestExportRoutes:
    """Functional tests for export routes"""

    def test_export_roadmap_pptx_success(self, client, auth_headers, category_with_features):
        """Test exporting roadmap to PowerPoint"""
        response = client.post('/api/export/roadmap/pptx', headers=auth_headers)

        assert response.status_code == 200
        assert response.content_type == 'application/vnd.openxmlformats-officedocument.presentationml.presentation'

        # Verify it's a valid PowerPoint file
        pptx_data = BytesIO(response.data)
        prs = Presentation(pptx_data)
        assert len(prs.slides) >= 1  # At least title slide

    def test_export_roadmap_pptx_no_dates(self, client, auth_headers, sample_category):
        """Test exporting roadmap when no features have dates"""
        # Create feature without release date
        from app.models import Feature
        from app import db

        feature = Feature(
            id='F-NO-DATE',
            title='No Date Feature',
            category_id=sample_category.id,
            release_date=None
        )
        db.session.add(feature)
        db.session.commit()

        response = client.post('/api/export/roadmap/pptx', headers=auth_headers)

        assert response.status_code == 400
        assert 'error' in response.json

    def test_export_prd_excel_success(self, client, auth_headers, category_with_features):
        """Test exporting PRD to Excel"""
        response = client.post('/api/export/prd/excel', headers=auth_headers)

        assert response.status_code == 200
        assert response.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        # Verify it's a valid Excel file with category sheets (no PRD Summary)
        excel_data = BytesIO(response.data)
        wb = load_workbook(excel_data)
        # Should have at least one category sheet
        assert len(wb.sheetnames) >= 1

    def test_export_prd_word_success(self, client, auth_headers, category_with_features):
        """Test exporting PRD to Word"""
        response = client.post('/api/export/prd/word', headers=auth_headers)

        assert response.status_code == 200
        assert response.content_type == 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'

        # Verify it's a valid Word file
        word_data = BytesIO(response.data)
        doc = Document(word_data)
        assert len(doc.paragraphs) > 0

    def test_export_prd_excel_no_categories(self, client, auth_headers):
        """Test exporting PRD when no categories exist"""
        response = client.post('/api/export/prd/excel', headers=auth_headers)

        assert response.status_code == 400
        assert 'error' in response.json

    def test_export_prd_word_no_categories(self, client, auth_headers):
        """Test exporting PRD when no categories exist"""
        response = client.post('/api/export/prd/word', headers=auth_headers)

        assert response.status_code == 400
        assert 'error' in response.json

    def test_export_unauthorized(self, client):
        """Test exports without authentication"""
        routes = [
            '/api/export/roadmap/pptx',
            '/api/export/prd/excel',
            '/api/export/prd/word'
        ]

        for route in routes:
            response = client.post(route)
            assert response.status_code == 401


@pytest.mark.functional
class TestImportRoutes:
    """Functional tests for import routes"""

    def test_import_prd_excel_success(self, client, auth_headers, category_with_features, sample_project):
        """Test importing PRD from Excel"""
        from app import db
        from app.models import Project

        # First export to get a valid Excel file
        export_response = client.post('/api/export/prd/excel', headers=auth_headers)
        excel_data = export_response.data

        # Clear database but keep project
        db.session.query(Feature).delete()
        db.session.query(Category).delete()
        db.session.commit()

        # Ensure default project exists for import
        if not Project.query.get(1):
            default_proj = Project(id=1, name='Default Project', description='For imports')
            db.session.add(default_proj)
            db.session.commit()

        # Import the Excel file
        response = client.post(
            '/api/import/prd/excel',
            headers=auth_headers,
            data={'file': (BytesIO(excel_data), 'test.xlsx')},
            content_type='multipart/form-data'
        )

        if response.status_code != 200:
            print(f"Import failed with status {response.status_code}")
            print(f"Response: {response.get_json()}")

        assert response.status_code == 200
        assert 'categories_imported' in response.json
        assert 'features_imported' in response.json

    def test_import_prd_excel_no_file(self, client, auth_headers):
        """Test importing without providing a file"""
        response = client.post(
            '/api/import/prd/excel',
            headers=auth_headers,
            content_type='multipart/form-data'
        )

        assert response.status_code == 400
        assert 'error' in response.json

    def test_import_prd_excel_invalid_file_type(self, client, auth_headers):
        """Test importing with invalid file type"""
        response = client.post(
            '/api/import/prd/excel',
            headers=auth_headers,
            data={'file': (BytesIO(b'not excel data'), 'test.txt')},
            content_type='multipart/form-data'
        )

        assert response.status_code == 400
        assert 'error' in response.json

    def test_import_unauthorized(self, client):
        """Test import without authentication"""
        response = client.post('/api/import/prd/excel')

        assert response.status_code == 401
